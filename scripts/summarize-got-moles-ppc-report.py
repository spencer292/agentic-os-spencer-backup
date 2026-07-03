#!/usr/bin/env python
"""Structured summary of the Got Moles PPC report."""
import openpyxl
from pathlib import Path

XLSX = Path.home() / "Downloads" / "Got Moles - PPC Growth YoY + SEO Progress and Fruits (2).xlsx"
wb = openpyxl.load_workbook(XLSX, data_only=True)

print("=" * 70)
print("PPC SPEND + LEADS (Sheet5)")
print("=" * 70)
ws = wb['Sheet5']
for row in ws.iter_rows(values_only=True):
    cells = [str(c) if c is not None else '' for c in row]
    if any(c.strip() for c in cells):
        print("  " + " | ".join(c for c in cells if c.strip()))

print("\n" + "=" * 70)
print("SEO TOTALS (Overview top)")
print("=" * 70)
ws = wb['Overview']
for r in range(1, 6):
    row = [ws.cell(r, col).value for col in range(1, 5)]
    cells = [str(c) if c is not None else '' for c in row]
    if any(c.strip() for c in cells):
        print("  " + " | ".join(c for c in cells if c.strip()))

print("\n" + "=" * 70)
print("RANK MOVEMENTS (Overview)")
print("=" * 70)

all_movers = []
current_section = "(unnamed)"
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    b = ws.cell(r, 2).value
    c_val = ws.cell(r, 3).value
    prev_val = ws.cell(r, 4).value

    if isinstance(a, str) and any(t.lower() in a.lower() for t in ['fruits', 'top', 'page', 'progress']):
        current_section = a.strip().replace('\n', ' ')
        continue

    if isinstance(b, str) and b.strip().lower() == 'keyword':
        continue

    if isinstance(b, str) and b.strip() and isinstance(c_val, (int, float)) and isinstance(prev_val, (int, float)):
        all_movers.append({
            'keyword': b.strip(),
            'current': int(c_val),
            'previous': int(prev_val),
            'improvement': int(prev_val) - int(c_val),
            'section': current_section,
        })

all_movers.sort(key=lambda x: -x['improvement'])

print(f"\nTotal movement entries: {len(all_movers)}")
improved = [m for m in all_movers if m['improvement'] > 0]
declined = [m for m in all_movers if m['improvement'] < 0]
static = [m for m in all_movers if m['improvement'] == 0]
print(f"Improved: {len(improved)}, Declined: {len(declined)}, Static: {len(static)}")

print("\n--- TOP 40 RANK IMPROVEMENTS ---")
for m in improved[:40]:
    kw = m['keyword'][:60]
    print(f"  +{m['improvement']:>3}  {m['previous']:>3} -> {m['current']:>3}  {kw}")

print("\n--- KEYWORDS NOW IN TOP 3 (sample) ---")
top3 = sorted([m for m in all_movers if m['current'] <= 3], key=lambda x: x['current'])
print(f"Total in top 3: {len(top3)}")
for m in top3[:40]:
    kw = m['keyword'][:60]
    print(f"  #{m['current']}   was #{m['previous']:>3}  +{m['improvement']:>3}  {kw}")

print("\n--- RANK DECLINES ---")
for m in declined[:30]:
    kw = m['keyword'][:60]
    print(f"  {m['improvement']:>4}  {m['previous']:>3} -> {m['current']:>3}  {kw}")

print("\n" + "=" * 70)
print("RANKINGS SHEET")
print("=" * 70)
ws = wb['Rankings']
print(f"Dimensions: {ws.max_row} rows x {ws.max_column} cols")
for r in range(1, min(40, ws.max_row + 1)):
    row = [ws.cell(r, col).value for col in range(1, min(8, ws.max_column + 1))]
    cells = [str(c) if c is not None else '' for c in row]
    if any(c.strip() for c in cells):
        print(f"  R{r}: " + " | ".join(c for c in cells if c.strip()))
