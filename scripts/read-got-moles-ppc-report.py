#!/usr/bin/env python
"""Extract all data from the Got Moles PPC report."""
import openpyxl
from pathlib import Path

XLSX = Path.home() / "Downloads" / "Got Moles - PPC Growth YoY + SEO Progress and Fruits (2).xlsx"

wb = openpyxl.load_workbook(XLSX, data_only=True)

print(f"Sheets in workbook: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*80}\nSHEET: {sheet_name}  ({ws.max_row} rows x {ws.max_column} cols)\n{'='*80}\n")

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Skip fully empty rows
        if not any(c is not None and str(c).strip() != '' for c in row):
            continue
        # Print row, truncating long cells
        cells = []
        for c in row:
            if c is None:
                cells.append('')
            elif isinstance(c, float):
                # format floats nicely
                if c == int(c):
                    cells.append(str(int(c)))
                else:
                    cells.append(f"{c:.4f}".rstrip('0').rstrip('.'))
            else:
                s = str(c).strip()
                if len(s) > 80:
                    s = s[:77] + '...'
                cells.append(s)
        print(f"  R{row_idx}: " + " | ".join(cells))
        if row_idx > 200:
            print(f"  ... [truncated at row 200, sheet has {ws.max_row} rows]")
            break
