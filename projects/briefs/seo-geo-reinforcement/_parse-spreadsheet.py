"""Parse the previous-agency spreadsheet to extract per-city paid+ranking data for
the 6 undecided cities: Algona, Fairwood, Lake Tapps, Medina, Centralia, Eatonville."""

import openpyxl
from pathlib import Path

path = Path(r'C:\Claude\agent-os\clients\got-moles\projects\briefs\website-rebuild-rebrand\Got Moles - PPC Growth YoY + SEO Progress and Fruits.xlsx')

wb = openpyxl.load_workbook(path, data_only=True)

print("=" * 80)
print("SHEETS IN WORKBOOK:")
print("=" * 80)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"  - {sheet_name} ({ws.max_row} rows × {ws.max_column} cols)")

print()

# Print headers + first 3 rows of every sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'=' * 80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'=' * 80}")
    for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True):
        print([str(c)[:40] if c is not None else None for c in row])

# Search for the 6 target cities across all sheets
TARGET_CITIES = ['algona', 'fairwood', 'lake tapps', 'lake-tapps', 'laketapps', 'medina', 'centralia', 'eatonville']

print(f"\n\n{'=' * 80}")
print(f"CITY SEARCH: {TARGET_CITIES}")
print(f"{'=' * 80}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Sheet: {sheet_name} ---")
    headers = [cell.value for cell in ws[1]]
    print(f"  Headers: {headers}")

    hits_by_city = {c: [] for c in TARGET_CITIES}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_text = ' '.join(str(c).lower() for c in row if c is not None)
        for city in TARGET_CITIES:
            if city in row_text:
                hits_by_city[city].append(row)

    for city, hits in hits_by_city.items():
        if hits:
            print(f"\n  {city.upper()} — {len(hits)} hits:")
            for h in hits[:15]:  # cap to 15 per city per sheet
                # print row with column labels
                pairs = [f"{hdr}={str(val)[:50]}" for hdr, val in zip(headers, h) if val is not None]
                print(f"    • {' | '.join(pairs)}")
